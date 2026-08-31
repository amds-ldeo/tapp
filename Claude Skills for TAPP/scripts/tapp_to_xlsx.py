#!/usr/bin/env python3
"""
tapp_to_xlsx.py
---------------
Convert a TAPP CSV file to a formatted xlsx file for sharing.

Usage:
    python scripts/tapp_to_xlsx.py <input.csv> [output.xlsx]

If output path is omitted, the xlsx is written alongside the CSV with the
same base name (e.g., LA-ICP-MS_TAPP_v6.csv → LA-ICP-MS_TAPP_v6.xlsx).

What this script does:
  - Reads the CSV (source of truth)
  - Applies tier color coding per the TAPP conventions
  - Styles group header rows (bold, fill color, no merged cells)
  - Sets column widths per conventions
  - Adds a Legends sheet with tier and mode column definitions
  - Never modifies the CSV

Column structure (columns A–I are fixed; I+1 onward is dynamic):
  A  Metadata Item
  B  Description / Purpose
  C  Procedure-Level Tier
  D  Analysis-Level Tier
  E  Data Type
  F  Example / Allowed Content
  G  Comments (short field-level qualifiers that are neither mode nor cardinality —
       instrument variant, signal/detector, conditional notes. Cardinality moved to
       column I under Rule 7; the old "Analyte-Specific" labels were retired with it.)
  H  Last Update
  I  Keyed By (Rule 7) — never blank on a content row; `(none)` for a scalar
  I+1 … I+n   Mode flag columns (one per mode defined in Phase 0)
  I+n+1     Sentinel column — header must be exactly "Literature Assessment";
             all data rows are empty. Marks the boundary between mode flag
             columns and literature assessment columns.
  I+n+2 …  Literature assessment columns (one per extracted procedure)

  Fallback: if no sentinel column is found, the script uses a length-based
  heuristic (header ≤ 25 chars → mode flag; longer → literature assessment).
  This preserves compatibility with TAPPs created before the sentinel convention.

Conventions applied (from references/conventions.md):
  Procedure-Level Tier colors:
    Basic    → bold, red text  (C00000)
    Advanced → bold, green text (375623)
    N/A      → bold, default text

  Analysis-Level Tier colors:
    Read-Only → bold, blue text   (0070C0)
    Editable  → bold, purple text (7030A0)
    Basic     → bold, red text    (C00000)
    Advanced  → bold, green text  (375623)
    (N/A is not a valid analysis-level tier)

  Group header rows: bold, light blue fill (D9E1F2), all columns styled
  Mode flag columns: Y → light green fill; N → light red fill
  Sentinel column: narrow width (4 chars), light grey fill (D9D9D9)
  All content cells: wrap text, top-aligned
"""

import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
TIER_STYLES = {
    # Procedure-Level Tier (col C) and Analysis-Level Tier (col D)
    'Basic':      {'font_color': 'C00000', 'bold': True},
    'Advanced':   {'font_color': '375623', 'bold': True},
    'Read-Only':  {'font_color': '0070C0', 'bold': True},
    'Editable':   {'font_color': '7030A0', 'bold': True},
    'N/A':        {'font_color': '000000', 'bold': True},
}

HEADER_FILL   = PatternFill('solid', fgColor='D9E1F2')  # light blue for group headers
MODE_Y_FILL   = PatternFill('solid', fgColor='C6EFCE')  # light green  Y
MODE_N_FILL   = PatternFill('solid', fgColor='FFC7CE')  # light red    N
LIT_FILL      = PatternFill('solid', fgColor='FFF2CC')  # light yellow for lit assessment headers
SENTINEL_FILL = PatternFill('solid', fgColor='D9D9D9')  # light grey for sentinel column

SENTINEL_HEADER = 'Literature Assessment'  # exact string used to mark the mode/lit boundary

WRAP_TOP = Alignment(wrap_text=True, vertical='top')
WRAP_CTR = Alignment(wrap_text=True, vertical='top', horizontal='center')

# ---------------------------------------------------------------------------
# Column widths (characters) — per conventions.md
# ---------------------------------------------------------------------------
COL_WIDTHS = {
    1: 38,   # A — Metadata Item
    2: 68,   # B — Description
    3: 15,   # C — Procedure-Level Tier
    4: 15,   # D — Analysis-Level Tier
    5: 20,   # E — Data Type
    6: 48,   # F — Example/Allowed Content
    7: 22,   # G — Comments
    8: 14,   # H — Last update
    9: 24,   # I — Keyed By (Rule 7)
    # J onward: set dynamically below
}
MODE_COL_WIDTH = 13
LIT_COL_WIDTH  = 30

# ---------------------------------------------------------------------------
# Detect structure from header row
# ---------------------------------------------------------------------------
def detect_structure(rows):
    """
    Returns (mode_cols, lit_cols, sentinel_col) where each is a list or int
    of 0-based column indices.

    Primary method: look for a column whose header is exactly SENTINEL_HEADER
    ('Literature Assessment'). Columns between H and the sentinel are mode flag
    columns; columns after the sentinel are literature assessment columns.

    Fallback (no sentinel found): use the old length-based heuristic — headers
    ≤ 25 chars with no newlines are treated as mode flag columns; longer headers
    are literature assessment columns. This preserves compatibility with TAPPs
    created before the sentinel convention.
    """
    header = rows[0]
    mode_cols    = []
    lit_cols     = []
    sentinel_col = None

    # Fixed columns: A=0 B=1 C=2 D=3 E=4 F=5 G=6 H=7
    # Mode flags start at index 8 (I)
    for i in range(9, len(header)):
        h = header[i].strip()
        if h == SENTINEL_HEADER:
            sentinel_col = i
            continue
        if sentinel_col is None:
            if h:
                mode_cols.append(i)
        else:
            if h:
                lit_cols.append(i)

    # Fallback: sentinel not present — use length heuristic
    if sentinel_col is None:
        mode_cols, lit_cols = [], []
        in_mode = True
        for i in range(9, len(header)):
            h = header[i].strip()
            if not h:
                continue
            if in_mode and '\n' not in h and len(h) <= 25:
                mode_cols.append(i)
            else:
                in_mode = False
                lit_cols.append(i)

    return mode_cols, lit_cols, sentinel_col


# ---------------------------------------------------------------------------
# Identify group header rows
# ---------------------------------------------------------------------------
def is_group_header(row):
    """True if this row is a group section header (all other cols empty/N)."""
    a = row[0].strip()
    if not a:
        return False
    # Group headers start with a digit and a dot: "1. ", "2. " etc.
    if len(a) >= 3 and a[0].isdigit() and a[1] == '.':
        return True
    return False


def is_blank_row(row):
    return all(v.strip() == '' for v in row)


# ---------------------------------------------------------------------------
# Apply tier formatting to a cell
# ---------------------------------------------------------------------------
def apply_tier(cell, value):
    style = TIER_STYLES.get(value.strip())
    if style:
        cell.font = Font(bold=style['bold'], color=style['font_color'])
    cell.alignment = WRAP_CTR


# ---------------------------------------------------------------------------
# Build Legends sheet
# ---------------------------------------------------------------------------
def build_legends(wb, mode_headers, keys_used=()):
    ws = wb.create_sheet('Legends')

    bold = Font(bold=True, size=11)
    header_font = Font(bold=True, size=12)

    def hdr(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = header_font
        c.alignment = WRAP_TOP

    def row_pair(row, col, label, definition, label_color=None):
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = Font(bold=True, color=label_color or '000000')
        lc.alignment = WRAP_TOP
        dc = ws.cell(row=row, column=col+1, value=definition)
        dc.alignment = WRAP_TOP

    r = 1
    hdr(r, 1, 'Procedure-Level Tier'); hdr(r, 2, 'Description')
    hdr(r, 3, 'Analysis-Level Tier'); hdr(r, 4, 'Description')
    r += 1

    proto_tiers = [
        ('Basic',    'C00000', 'Mandatory for procedure registration. Must be provided to register a valid procedure.'),
        ('Advanced', '375623', 'Optional for procedure registration. Strongly recommended but not required.'),
        ('N/A',      '000000', 'Not applicable at procedure level. This field captures analysis-level information only.'),
    ]
    analysis_tiers = [
        ('Read-Only', '0070C0', 'Directly imported from the registered procedure; cannot be changed by the analyst. Changing this value means running a different procedure. Also used for fields relevant only at procedure level — the value is inherited from the procedure record and shown read-only in the analysis form.'),
        ('Editable',  '7030A0', 'Imported from the registered procedure but may be adjusted within procedure-defined bounds (e.g., daily tuning, minor software updates). The procedure registers the target or typical value; the analyst confirms or adjusts it. Cannot be left void if the procedure-level tier is Basic.'),
        ('Basic',     'C00000', 'Mandatory user input at analysis time. Value comes from the analysis itself and cannot be pre-specified in the procedure.'),
        ('Advanced',  '375623', 'Optional user input at analysis time. Recommended for complete and reproducible documentation.'),
    ]

    for i, (label, color, defn) in enumerate(proto_tiers):
        row_pair(r+i, 1, label, defn, color)
    for i, (label, color, defn) in enumerate(analysis_tiers):
        row_pair(r+i, 3, label, defn, color)

    r += max(len(proto_tiers), len(analysis_tiers)) + 2

    hdr(r, 1, 'Mode Column'); hdr(r, 2, 'Definition')
    r += 1
    for mode_name in mode_headers:
        ws.cell(row=r, column=1, value=mode_name).font = bold
        ws.cell(row=r, column=1).alignment = WRAP_TOP
        ws.cell(row=r, column=2).alignment = WRAP_TOP
        r += 1

    row_pair(r,   1, 'Y', 'This field applies to this analytical mode and should be reported.')
    row_pair(r+1, 1, 'N', 'This field does not apply to this analytical mode.')


    # --- Data Type table -----------------------------------------------------
    # Added 2026-08-30 with the two-type scheme. `Other: specify` was removed from 226
    # Column F cells: on a closed `Controlled list` it contradicted the type, and on a
    # `Controlled list / Text` it asked for the wrong thing. The guidance it used to
    # carry inline lives here instead — stated once, where it cannot drift out of sync
    # with the type it describes.
    r += 3
    hdr(r, 1, 'Data Type'); hdr(r, 2, 'What to enter')
    r += 1
    TYPE_DEFS = [
        ('Controlled list',
         'Choose one of the values listed in "Example / Allowed Content". The list is '
         'CLOSED — it is meant to cover every case. If your value genuinely is not there, '
         'that is a gap in the list worth reporting, not a reason to write your own.'),
        ('Controlled list / Text',
         'Choose a listed value AND qualify it in the same cell — the detail is expected, '
         'not optional. Typically what was corrected, which analytes or phases it applied '
         'to, the method or equation used, or the source it came from. An answer the list '
         'cannot express is also valid here.'),
        ('Numeric (unit)',
         'A number in the unit named in brackets. Enter the number only — the unit is '
         'fixed by the field.'),
        ('Numeric + unit',
         'A number AND its unit, because the unit varies between procedures. Write both, '
         "e.g. '50 us' or '0.5 s'."),
        ('Text (free)', 'Free text. No controlled vocabulary applies.'),
        ('N/A | None',
         'Offered by every controlled list. "N/A" means the field does not apply to this '
         'procedure; "None" means it applies but nothing was used. Prefer either over '
         'leaving the cell blank.'),
    ]
    for label, defn in TYPE_DEFS:
        row_pair(r, 1, label, defn)
        r += 1

    r += 3
    hdr(r, 1, 'Keyed By (Rule 7)'); hdr(r, 2, 'What the field\'s value repeats over')
    r += 1
    KEY_DEFS = {
        '(none)': 'Scalar — one value per procedure or analysis. The default.',
        'sampling unit': 'One value per subdivision of the physical sample that carries its own row — grain, spot, aliquot, phase, sub-volume.',
        'reported property': 'One value per reported quantity or nominal property, at any point in the chain — ratios and dates alike, plus their uncertainties.',
        'channel': "One value per position on the instrument's selection axis — the address, not the signal. Mass, cup, line + crystal, energy-loss edge, wavenumber.",
        'analyte': 'One value per chemical species determined, at whatever granularity the procedure determines it.',
        'standard': 'One value per reference material or reference database entry.',
        'conversion': 'One value per correction or calculation step, where it cannot be attributed to a single reported property.',
        'model component': 'One value per component of a fitted decomposition of the signal.',
        'acquisition pass': 'One value per pass over the sample with its own instrument settings.',
        'preparation step': 'One value per stage in sample preparation.',
        'background position': 'One value per off-peak background position.',
    }
    for k in keys_used:
        # `defines: <domain> per <key>` carries two roles and must be split on `per`
        # first — the generic `defines:` branch below would otherwise treat the whole
        # tail as one domain name and drop the key's definition from the legend.
        m_per = re.match(r'^defines:\s*(.+?)\s+per\s+(.+)$', k)
        if m_per:
            dom, key = m_per.group(1).strip(), m_per.group(2).strip()
            defn = ('ENUMERATES the %s domain — the header of the child table, not a column '
                    'in it — and does so once per %s. %s %s'
                    % (dom, key, KEY_DEFS.get(dom, ''), KEY_DEFS.get(key, ''))).strip()
            ws.cell(row=r, column=1, value=k).font = bold
            ws.cell(row=r, column=1).alignment = WRAP_TOP
            ws.cell(row=r, column=2, value=defn).alignment = WRAP_TOP
            r += 1
            continue
        base = k.replace('defines: ', '').replace('pair: ', '')
        parts = [p.strip() for p in re.split(r'\s*(?:>|x)\s*', base) if p.strip()]
        defn = ' + '.join(KEY_DEFS.get(p, '') for p in parts).strip(' +')
        if k.startswith('defines: '):
            defn = 'ENUMERATES the key domain rather than being keyed by it — the header of the child table, not a column in it. ' + defn
        elif k.startswith('pair: '):
            defn = 'One value per unordered PAIR. ' + defn
        elif ' x ' in k:
            defn = 'Cross-product, read as "for each %s, one value per %s". ' % (parts[0], parts[-1]) + defn
        elif ' > ' in k:
            defn = 'Nested — %s exists only within %s. ' % (parts[-1], parts[0]) + defn
        ws.cell(row=r, column=1, value=k).font = bold
        ws.cell(row=r, column=1).alignment = WRAP_TOP
        ws.cell(row=r, column=2, value=defn).alignment = WRAP_TOP
        r += 1

    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 75
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 55


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------
def convert(csv_path, xlsx_path):
    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))

    if not rows:
        print("ERROR: CSV is empty.")
        sys.exit(1)

    mode_cols, lit_cols, sentinel_col = detect_structure(rows)
    mode_headers = [rows[0][i] for i in mode_cols]
    # Rule 7 — Legends Table 4 lists only the keys this TAPP actually uses.
    _seen, keys_used = set(), []
    for _r in rows[1:]:
        _k = _r[8].strip() if len(_r) > 8 else ''
        if _k and _k not in _seen:
            _seen.add(_k); keys_used.append(_k)
    keys_used.sort(key=lambda k: (k != '(none)', k))
    n_cols = len(rows[0])

    print(f"  Rows: {len(rows)}")
    print(f"  Columns: {n_cols}")
    print(f"  Mode flag columns ({len(mode_cols)}): {mode_headers}")
    print(f"  Sentinel column: {'col ' + str(sentinel_col) if sentinel_col is not None else 'not found (using fallback heuristic)'}")
    print(f"  Literature assessment columns: {len(lit_cols)}")

    wb = Workbook()
    ws = wb.active
    ws.title = 'TAPP'

    # Freeze panes: freeze row 1 (header) and cols A-H (8 fixed cols)
    ws.freeze_panes = ws.cell(row=2, column=9)

    for ri, row in enumerate(rows, start=1):
        a_val = row[0].strip() if row else ''

        if is_blank_row(row):
            # Blank separator row — write nothing, set small height
            ws.row_dimensions[ri].height = 8
            continue

        group_hdr = is_group_header(row)

        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val if val.strip() else None)
            cell.alignment = WRAP_TOP

            if group_hdr:
                # Style all cells in header row
                cell.font = Font(bold=True, size=11)
                cell.fill = HEADER_FILL
                cell.alignment = WRAP_TOP
                continue

            col_idx = ci - 1  # 0-based

            # Procedure-Level Tier (col C = index 2)
            if col_idx == 2 and val.strip():
                apply_tier(cell, val.strip())

            # Analysis-Level Tier (col D = index 3)
            elif col_idx == 3 and val.strip():
                apply_tier(cell, val.strip())

            # Mode flag columns
            elif col_idx in mode_cols:
                v = val.strip().upper()
                if v == 'Y':
                    cell.fill = MODE_Y_FILL
                elif v == 'N':
                    cell.fill = MODE_N_FILL
                cell.alignment = WRAP_CTR

            # Sentinel column — style all rows with grey fill; bold header
            elif col_idx == sentinel_col:
                cell.fill = SENTINEL_FILL
                if ri == 1:
                    cell.font = Font(bold=True, size=10)
                cell.alignment = WRAP_CTR

            # Literature assessment header row (row 1)
            elif col_idx in lit_cols and ri == 1:
                cell.font = Font(bold=True, size=10)
                cell.fill = LIT_FILL
                cell.alignment = WRAP_TOP

    # --- Column widths ---
    for col_idx_1based, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx_1based)].width = width

    # Mode flag columns
    for col_idx in mode_cols:
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = MODE_COL_WIDTH

    # Sentinel column — narrow
    if sentinel_col is not None:
        ws.column_dimensions[get_column_letter(sentinel_col + 1)].width = 4

    # Literature assessment columns
    for col_idx in lit_cols:
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = LIT_COL_WIDTH

    # --- Legends sheet ---
    build_legends(wb, mode_headers, keys_used)

    wb.save(xlsx_path)
    print(f"\nSaved: {xlsx_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    csv_path = Path(sys.argv[1])
    if not csv_path.exists():
        print(f"ERROR: File not found: {csv_path}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        xlsx_path = Path(sys.argv[2])
    else:
        xlsx_path = csv_path.with_suffix('.xlsx')

    print(f"Converting: {csv_path} → {xlsx_path}")
    convert(str(csv_path), str(xlsx_path))
